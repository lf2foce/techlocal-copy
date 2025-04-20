from sqlalchemy.orm import Session
from database.models import ContentPost, Theme, Campaign
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_excel_sheet(worksheet):
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Format headers
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Format data cells and adjust column width
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # Apply borders and center alignment to all cells
        for cell in column:
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            # Calculate maximum length for column width
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Adjust column width (with a maximum width to prevent too wide columns)
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def export_to_excel(db: Session) -> tuple[BytesIO, str]:
    # Query all data
    posts = db.query(ContentPost).all()
    campaigns = db.query(Campaign).all()
    themes = db.query(Theme).all()
    
    # Convert to pandas DataFrames with formatted data
    posts_data = [{
        'ID': post.id,
        'Campaign ID': post.campaign_id,
        'Theme ID': post.theme_id,
        'Content': post.content,
        'Status': post.status.upper() if post.status else '',
        'Created At': post.created_at.strftime('%Y-%m-%d %H:%M') if post.created_at else '',
        'Posted At': post.posted_at.strftime('%Y-%m-%d') if post.posted_at else '',
        'Image Status': post.image_status.upper() if post.image_status else ''
    } for post in posts]
    
    campaigns_data = [{
        'ID': campaign.id,
        'Title': campaign.title,
        'Repeat Days': campaign.repeat_every_days,
        'Target Customer': campaign.target_customer,
        'Insight': campaign.insight,
        'Description': campaign.description,
        'Status': campaign.status.upper() if campaign.status else '',
        'Start Date': campaign.start_date.strftime('%Y-%m-%d') if campaign.start_date else '',
        'Active': 'Yes' if campaign.is_active else 'No',
        'Current Step': campaign.current_step
    } for campaign in campaigns]
    
    themes_data = [{
        'ID': theme.id,
        'Campaign ID': theme.campaign_id,
        'Title': theme.title,
        'Story': theme.story,
        'Selected': 'Yes' if theme.is_selected else 'No',
        'Status': theme.status.upper() if theme.status else '',
        'Post Status': theme.post_status.upper() if theme.post_status else '',
        'Created At': theme.created_at.strftime('%Y-%m-%d %H:%M') if theme.created_at else ''
    } for theme in themes]
    
    # Extract image data from posts
    images_data = []
    for post in posts:
        try:
            if not post.images:
                continue

            # Handle different JSONB structures
            images_dict = {}
            if isinstance(post.images, dict):
                images_dict = post.images
            elif isinstance(post.images, str):
                try:
                    import json
                    images_dict = json.loads(post.images)
                except json.JSONDecodeError:
                    continue
            else:
                continue
            
            # Get images list from various possible structures
            image_list = []
            if 'images' in images_dict and isinstance(images_dict['images'], list):
                image_list = images_dict['images']
            elif 'data' in images_dict:
                if isinstance(images_dict['data'], list):
                    image_list = images_dict['data']
                elif isinstance(images_dict['data'], dict):
                    image_list = [images_dict['data']]
            elif 'url' in images_dict:
                image_list = [images_dict]

            # Process each image
            for image in image_list:
                if not isinstance(image, dict):
                    continue
                
                # Extract image data with proper type checking
                image_url = str(image.get('url', ''))
                image_prompt = str(image.get('prompt', ''))
                image_order = str(image.get('order', '')) if image.get('order') is not None else ''
                image_selected = bool(image.get('isSelected', False))
                image_status = str(image.get('status', '')).upper() if image.get('status') else ''
                
                images_data.append({
                    'Post ID': post.id,
                    'Campaign ID': post.campaign_id,
                    'Image URL': image_url,
                    'Prompt': image_prompt,
                    'Order': image_order,
                    'Selected': 'Yes' if image_selected else 'No',
                    'Status': image_status
                })
        except Exception as e:
            # Log the error but continue processing other posts
            print(f"Error processing images for post {post.id}: {str(e)}")
            continue
    
    # Create DataFrames
    posts_df = pd.DataFrame(posts_data)
    campaigns_df = pd.DataFrame(campaigns_data)
    themes_df = pd.DataFrame(themes_data)
    images_df = pd.DataFrame(images_data)
    
    # Create buffer and write Excel file
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Write DataFrames to Excel
        posts_df.to_excel(writer, sheet_name='Posts', index=False)
        campaigns_df.to_excel(writer, sheet_name='Campaigns', index=False)
        themes_df.to_excel(writer, sheet_name='Themes', index=False)
        images_df.to_excel(writer, sheet_name='Images', index=False)
        
        # Get workbook and format each sheet
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            format_excel_sheet(workbook[sheet_name])
    
    # Reset buffer position
    buffer.seek(0)
    
    return buffer, 'social_media_data.xlsx'